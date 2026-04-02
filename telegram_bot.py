"""
Dryft — Telegram Bot Interface

Routes messages through the Dryft proxy (context injection + signal
detection + ecological cycle) and returns the enriched response.
Telegram is the prompt terminal. The proxy is the brain.

Voice messages transcribed via Groq Whisper, responses synthesized
via edge-tts. Text always works; voice degrades gracefully if
dependencies are missing.

Commands:
    /status  — herd health summary
    /save    — force-save current state

Usage:
    export TELEGRAM_BOT_TOKEN=<your-token>
    export ANTHROPIC_API_KEY=<your-key>
    export GROQ_API_KEY=<your-key>        # optional, enables voice input
    export TELEGRAM_USER_ID=<your-id>     # required, restricts to owner
    python telegram_bot.py
"""

import asyncio
import base64
import functools
import json
import os
import sys
import tempfile
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import anthropic
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from proxy import DryftProxy

proxy = DryftProxy()

ANTHROPIC_BILLING_URL = "https://console.anthropic.com/settings/billing"

# Voice handlers (optional)
stt = None
stt_fallback = None
tts = None
try:
    from voice_handler import SpeechToText, AnthropicSTT, TextToSpeech, VoiceError, GroqLimitError
    stt = SpeechToText()
    stt_fallback = AnthropicSTT()
    tts = TextToSpeech()
    print("Voice handlers initialized (Groq Whisper + Anthropic fallback + edge-tts).")
except Exception as e:
    print(f"Voice not available: {e}")
    print("Voice messages will not be supported. Text still works.")
    try:
        from voice_handler import VoiceError, GroqLimitError
    except ImportError:
        class VoiceError(Exception):
            pass
        class GroqLimitError(VoiceError):
            pass

# Owner guard
OWNER_ID = int(os.environ.get("TELEGRAM_USER_ID", "0"))

# Input limits
MAX_TEXT_LENGTH = 4000       # characters — bounds API cost and memory extraction
MAX_VOICE_FILE_SIZE = 20_000_000  # 20 MB
MAX_IMAGE_FILE_SIZE = 10_000_000  # 10 MB — keeps vision API cost reasonable
MAX_DOCUMENT_FILE_SIZE = 10_000_000  # 10 MB
MAX_EXTRACTED_TEXT = 8000    # characters — truncate file extractions before feeding to proxy

# Vision model for image description
VISION_MODEL = "claude-sonnet-4-20250514"


def owner_only(func):
    """Silently ignore messages from non-owner users."""
    @functools.wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if OWNER_ID and update.effective_user.id != OWNER_ID:
            return
        return await func(update, context)
    return wrapper


async def _send_text(update: Update, text: str):
    """Send text reply, splitting into chunks if needed."""
    if len(text) <= 4096:
        await update.message.reply_text(text)
    else:
        for i in range(0, len(text), 4096):
            await update.message.reply_text(text[i:i + 4096])



def _get_last_morning_message_id() -> int:
    """Get the stored morning message ID from herd_meta.json."""
    meta_path = os.path.join(proxy.state_dir, "herd_meta.json")
    if os.path.exists(meta_path):
        with open(meta_path, "r") as f:
            meta = json.load(f)
        return meta.get("last_morning_message_id", 0)
    return 0


def _is_morning_feedback(update: Update) -> bool:
    """Check if this message is a reply to the bot's morning message."""
    if not update.message.reply_to_message:
        return False
    reply_msg_id = update.message.reply_to_message.message_id
    morning_msg_id = _get_last_morning_message_id()
    return morning_msg_id > 0 and reply_msg_id == morning_msg_id


def _write_morning_preference(preference_text: str):
    """Write a morning message preference to the grass layer."""
    from herd_engine import GrassEntry
    if not proxy.engine.grass_layer:
        return
    pref_id = f"morning-pref-{len(proxy.engine.grass_layer.entries)}"
    entry = GrassEntry({
        "id": pref_id,
        "content": preference_text,
        "source": "direct",
        "keywords": ["morning", "preference"],
    })
    proxy.engine.grass_layer.entries[pref_id] = entry
    proxy.save_state()
    print(f"Morning preference saved: {pref_id} -> {preference_text[:80]}")


@owner_only
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Route text message through Dryft proxy."""
    user_message = update.message.text
    chat_id = update.effective_chat.id

    if len(user_message) > MAX_TEXT_LENGTH:
        await update.message.reply_text(
            f"Message too long ({len(user_message)} chars). Limit is {MAX_TEXT_LENGTH}."
        )
        return

    # Check for morning message feedback
    if _is_morning_feedback(update):
        _write_morning_preference(user_message)
        await update.message.reply_text("Got it, I'll adjust future morning messages accordingly.")
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    try:
        result = proxy.process(user_message)
    except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
        if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or e.status_code == 429:
            await update.message.reply_text(
                f"Anthropic API credits exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}"
            )
            return
        raise
    except Exception as e:
        print(f"Message handler error: {e}")
        await update.message.reply_text("Something went wrong. Try again.")
        return

    response = result["response"]

    if response:
        await _send_text(update, response)
    else:
        await update.message.reply_text("No response generated.")


@owner_only
async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle voice message: transcribe, process, respond with text + voice."""
    chat_id = update.effective_chat.id

    if not stt:
        await update.message.reply_text("Voice input not available. Send text instead.")
        return

    # Check voice file size before downloading
    voice = update.message.voice
    if voice.file_size and voice.file_size > MAX_VOICE_FILE_SIZE:
        await update.message.reply_text(
            f"Voice message too large ({voice.file_size // 1_000_000}MB). Limit is 20MB."
        )
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    # Download the voice file
    voice_file = await context.bot.get_file(voice.file_id)

    tmp_in = tempfile.NamedTemporaryFile(suffix=".ogg", delete=False)
    tmp_in.close()

    try:
        await voice_file.download_to_drive(tmp_in.name)

        # Transcribe: try Groq (free), fall back to Anthropic (paid)
        groq_exhausted = False
        try:
            transcribed = await asyncio.to_thread(stt.transcribe, tmp_in.name)
        except GroqLimitError:
            groq_exhausted = True
            if stt_fallback:
                await update.message.reply_text("Groq free voice credits used up. Switching to paid voice via Anthropic.")
                try:
                    transcribed = await asyncio.to_thread(stt_fallback.transcribe, tmp_in.name)
                except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
                    if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or e.status_code == 429:
                        await update.message.reply_text(
                            f"Anthropic API credits also exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}"
                        )
                        return
                    raise
            else:
                await update.message.reply_text("Groq free voice credits used up. Send text instead.")
                return

        if not transcribed or not transcribed.strip():
            await update.message.reply_text("Could not transcribe audio. Try again or send text.")
            return

        # Process through Dryft proxy
        try:
            result = proxy.process(transcribed)
        except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
            if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or e.status_code == 429:
                await update.message.reply_text(
                    f"Anthropic API credits exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}"
                )
                return
            raise

        response = result["response"]

        if not response:
            await update.message.reply_text("No response generated.")
            return

        # Send text reply (with transcription for verification)
        source_label = "Heard via Anthropic" if groq_exhausted else "Heard"
        text_reply = f"[{source_label}: {transcribed}]\n\n{response}"
        await _send_text(update, text_reply)

        # Generate and send voice reply
        if tts:
            tmp_out = tempfile.NamedTemporaryFile(suffix=".mp3", delete=False)
            tmp_out.close()
            try:
                await tts.synthesize(response, tmp_out.name)
                with open(tmp_out.name, "rb") as audio:
                    await update.message.reply_voice(voice=audio)
            except VoiceError:
                pass  # Text already sent; voice is a bonus
            finally:
                if os.path.exists(tmp_out.name):
                    os.unlink(tmp_out.name)

    except VoiceError as e:
        await update.message.reply_text(f"Voice processing failed: {e}\nSend text instead.")
    except Exception as e:
        print(f"Voice handler error: {e}")  # Log full error server-side
        await update.message.reply_text("Something went wrong processing your voice message. Try again or send text.")
    finally:
        if os.path.exists(tmp_in.name):
            os.unlink(tmp_in.name)


def _describe_image(image_bytes: bytes, media_type: str = "image/jpeg") -> str:
    """Send image to Claude vision API for objective description."""
    client = anthropic.Anthropic()
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    response = client.messages.create(
        model=VISION_MODEL,
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": b64,
                    },
                },
                {
                    "type": "text",
                    "text": (
                        "Describe this image objectively and concisely. "
                        "Focus on what is shown: content, text, data, objects, people, places. "
                        "Do not speculate about why the user shared it. "
                        "If the image contains a document, spreadsheet, or screenshot, "
                        "prioritize extracting the readable text and data."
                    ),
                },
            ],
        }],
    )
    return response.content[0].text


@owner_only
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo messages: describe via vision API, route through memory pipeline."""
    chat_id = update.effective_chat.id

    # Grab the largest resolution
    photo = update.message.photo[-1]
    if photo.file_size and photo.file_size > MAX_IMAGE_FILE_SIZE:
        await update.message.reply_text(
            f"Image too large ({photo.file_size // 1_000_000}MB). Limit is {MAX_IMAGE_FILE_SIZE // 1_000_000}MB."
        )
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    try:
        photo_file = await context.bot.get_file(photo.file_id)
        image_bytes = await photo_file.download_as_bytearray()

        # Get vision description
        description = await asyncio.to_thread(_describe_image, bytes(image_bytes))

        # Build proxy input with caption if present
        caption = update.message.caption or ""
        if caption:
            proxy_input = f'User shared an image with caption "{caption}": {description}'
        else:
            proxy_input = f"User shared an image: {description}"

        # Truncate if needed
        if len(proxy_input) > MAX_EXTRACTED_TEXT:
            proxy_input = proxy_input[:MAX_EXTRACTED_TEXT] + "... [truncated]"

        # Route through memory pipeline
        result = proxy.process(proxy_input)
        response = result["response"]

        if response:
            await _send_text(update, response)
        else:
            await update.message.reply_text("No response generated.")

    except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
        if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or getattr(e, 'status_code', 0) == 429:
            await update.message.reply_text(
                f"Anthropic API credits exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}"
            )
            return
        raise
    except Exception as e:
        print(f"Photo handler error: {e}")
        await update.message.reply_text("Something went wrong processing your image. Try again or send text.")


# ── File extraction helpers ──────────────────────────────────────────────────

# Image MIME types that should route through vision instead of text extraction
IMAGE_MIME_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp"}

# Supported document MIME types and their labels
SUPPORTED_DOCUMENT_TYPES = {
    "application/pdf": "PDF",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "Word document",
    "application/msword": "Word document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "Excel spreadsheet",
    "application/vnd.ms-excel": "Excel spreadsheet",
    "text/plain": "text file",
    "text/markdown": "Markdown file",
    "text/csv": "CSV file",
    "application/csv": "CSV file",
}


def _extract_text(file_path: str, mime_type: str) -> str:
    """Extract text content from a file based on MIME type."""
    if mime_type == "application/pdf":
        try:
            import PyPDF2
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                pages = [page.extract_text() or "" for page in reader.pages]
                return "\n".join(pages).strip()
        except Exception as e:
            return f"[PDF extraction failed: {e}]"

    if mime_type in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"):
        try:
            import docx
            doc = docx.Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs).strip()
        except Exception as e:
            return f"[Word extraction failed: {e}]"

    if mime_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
            wb.close()
            return "\n".join(lines).strip()
        except Exception as e:
            return f"[Excel extraction failed: {e}]"

    if mime_type in ("text/plain", "text/markdown", "text/csv", "application/csv"):
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read().strip()
        except Exception as e:
            return f"[Text extraction failed: {e}]"

    return ""


@owner_only
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document attachments: extract text or route images through vision."""
    chat_id = update.effective_chat.id
    doc = update.message.document

    if doc.file_size and doc.file_size > MAX_DOCUMENT_FILE_SIZE:
        await update.message.reply_text(
            f"File too large ({doc.file_size // 1_000_000}MB). Limit is {MAX_DOCUMENT_FILE_SIZE // 1_000_000}MB."
        )
        return

    mime_type = doc.mime_type or ""

    # Images sent as documents route through vision
    if mime_type in IMAGE_MIME_TYPES:
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")
        try:
            file_obj = await context.bot.get_file(doc.file_id)
            image_bytes = await file_obj.download_as_bytearray()
            description = await asyncio.to_thread(_describe_image, bytes(image_bytes), mime_type)
            caption = update.message.caption or ""
            if caption:
                proxy_input = f'User shared an image with caption "{caption}": {description}'
            else:
                proxy_input = f"User shared an image: {description}"
            if len(proxy_input) > MAX_EXTRACTED_TEXT:
                proxy_input = proxy_input[:MAX_EXTRACTED_TEXT] + "... [truncated]"
            result = proxy.process(proxy_input)
            response = result["response"]
            if response:
                await _send_text(update, response)
            else:
                await update.message.reply_text("No response generated.")
        except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
            if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or getattr(e, 'status_code', 0) == 429:
                await update.message.reply_text(f"Anthropic API credits exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}")
                return
            raise
        except Exception as e:
            print(f"Document image handler error: {e}")
            await update.message.reply_text("Something went wrong processing your image.")
        return

    # Text-based document extraction
    if mime_type not in SUPPORTED_DOCUMENT_TYPES:
        file_name = doc.file_name or "unknown"
        await update.message.reply_text(
            f"Unsupported file type: {file_name} ({mime_type or 'unknown type'})\n"
            f"Supported: PDF, Word, Excel, Markdown, text, CSV."
        )
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    file_label = SUPPORTED_DOCUMENT_TYPES[mime_type]
    suffix = os.path.splitext(doc.file_name or "file")[1] or ".tmp"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    tmp.close()

    try:
        file_obj = await context.bot.get_file(doc.file_id)
        await file_obj.download_to_drive(tmp.name)

        extracted = await asyncio.to_thread(_extract_text, tmp.name, mime_type)

        if not extracted or extracted.startswith("["):
            await update.message.reply_text(f"Could not extract text from {file_label}.")
            return

        # Truncate
        if len(extracted) > MAX_EXTRACTED_TEXT:
            extracted = extracted[:MAX_EXTRACTED_TEXT] + "... [truncated]"

        caption = update.message.caption or ""
        if caption:
            proxy_input = f'User shared a {file_label} with caption "{caption}" containing:\n{extracted}'
        else:
            proxy_input = f"User shared a {file_label} containing:\n{extracted}"

        result = proxy.process(proxy_input)
        response = result["response"]

        if response:
            await _send_text(update, response)
        else:
            await update.message.reply_text("No response generated.")

    except (anthropic.RateLimitError, anthropic.APIStatusError) as e:
        if "credit" in str(e).lower() or "billing" in str(e).lower() or "budget" in str(e).lower() or getattr(e, 'status_code', 0) == 429:
            await update.message.reply_text(f"Anthropic API credits exhausted.\nAdd credits: {ANTHROPIC_BILLING_URL}")
            return
        raise
    except Exception as e:
        print(f"Document handler error: {e}")
        await update.message.reply_text(f"Something went wrong processing your {file_label}. Try again.")
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


async def whoami_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Return the sender's Telegram user ID. No owner guard so you can discover your ID."""
    user = update.effective_user
    await update.message.reply_text(f"Your Telegram user ID: {user.id}")


@owner_only
async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Return herd health summary."""
    s = proxy.status()
    herd = s["herd"]

    all_memories = proxy.engine.herd_status()
    active = [m for m in all_memories if m["status"] == "active"]
    top5 = active[:5]

    lines = [
        f"Model: {proxy.model}",
        f"Herd: {herd['active']} active, {herd['culled']} culled",
        f"Foundational: {s['foundational']} memories",
        f"Grass: {s['grass']} entries",
        f"Staging: {s['staging']} signals incubating",
        f"Queries: {s['queries_processed']} processed, {s['silent_queries']} silent",
        f"Threshold: {s['threshold']}",
        f"History: {len(proxy.conversation_history)} messages",
        "",
        "Top 5 by fitness:",
    ]
    for m in top5:
        lines.append(f"  {m['id']}: {m['fitness_score']:.3f}")

    bonds = proxy.engine.get_bonds()
    bond_summary = {}
    for b in bonds:
        bond_summary[b["status"]] = bond_summary.get(b["status"], 0) + 1
    bond_parts = [f"{count} {status}" for status, count in sorted(bond_summary.items())]
    lines.append(f"\nBonds: {len(bonds)} total ({', '.join(bond_parts)})" if bond_parts else "\nBonds: 0")

    # Suspicious death watchlist
    unseen = proxy.get_watchlist_unseen()
    if unseen > 0:
        lines.append(f"\nSuspicious culls: {unseen} since last check")
        proxy.clear_watchlist_unseen()

    await update.message.reply_text("\n".join(lines))


@owner_only
async def save_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Force-save current state."""
    proxy.save_state()
    await update.message.reply_text("State saved.")


@owner_only
async def flag_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Flag the previous bot response for review. Logs query, response, and activated memories."""
    # Need at least 2 entries in history (user + assistant)
    if len(proxy.conversation_history) < 2:
        await update.message.reply_text("Nothing to flag. Send a message first, then /flag if the response is wrong.")
        return

    # Get the last user query and bot response from conversation history
    last_assistant = None
    last_user = None
    for msg in reversed(proxy.conversation_history):
        if msg["role"] == "assistant" and last_assistant is None:
            last_assistant = msg["content"]
        elif msg["role"] == "user" and last_assistant is not None:
            last_user = msg["content"]
            break

    if not last_user or not last_assistant:
        await update.message.reply_text("Nothing to flag. Send a message first, then /flag if the response is wrong.")
        return

    # Build flag entry
    flag_entry = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "user_query": last_user,
        "bot_response": last_assistant,
        "activated_memory_ids": list(proxy.last_activation),
        "status": "pending",
    }

    # Append to flagged_responses.json (append-only JSON array)
    flag_path = os.path.join(proxy.state_dir, "flagged_responses.json")
    if os.path.exists(flag_path):
        with open(flag_path, "r") as f:
            flags = json.load(f)
    else:
        flags = []

    flags.append(flag_entry)

    os.makedirs(proxy.state_dir, exist_ok=True)
    with open(flag_path, "w") as f:
        json.dump(flags, f, indent=2)

    n_memories = len(flag_entry["activated_memory_ids"])
    await update.message.reply_text(f"Flagged. {n_memories} activated memories logged for review.")


# Model-switching commands
MODEL_IDS = {
    "haiku": "claude-haiku-4-5-20251001",
    "sonnet": "claude-sonnet-4-20250514",
    "opus": "claude-opus-4-20250514",
}


@owner_only
async def model_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Switch the conversation model. Usage: /haiku, /sonnet, /opus"""
    command = update.message.text.lstrip("/").split()[0].lower()
    model_id = MODEL_IDS.get(command)
    if model_id:
        proxy.model = model_id
        await update.message.reply_text(f"Switched to {command} ({model_id}).")


@owner_only
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show available commands."""
    lines = [
        "Dryft Commands",
        "",
        "/status — Herd health, top memories, bond summary, watchlist alerts",
        "/save — Force-save all state to disk",
        "/flag — Flag the last response as wrong (logged for review)",
        "/haiku — Switch to Haiku (fast, cheap)",
        "/sonnet — Switch to Sonnet (default, balanced)",
        "/opus — Switch to Opus (strongest, expensive)",
        "/help — This message",
        "",
        "Tips",
        "",
        "Say \"that's wrong\" or \"no, actually...\" to correct a bad memory.",
        "Reply to a morning message to adjust future briefings.",
        "Send photos, PDFs, Word docs, Excel files, or voice messages.",
        "Paste a Google Sheets URL to read sheet data.",
        "Ask about weather to get local conditions.",
    ]
    await update.message.reply_text("\n".join(lines))


def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        print("ERROR: Set TELEGRAM_BOT_TOKEN environment variable.")
        print("  export TELEGRAM_BOT_TOKEN=$(powershell -Command \"[Environment]::GetEnvironmentVariable('TELEGRAM_BOT_TOKEN', 'User')\" | tr -d '\\r\\n')")
        sys.exit(1)

    if not OWNER_ID:
        print("ERROR: TELEGRAM_USER_ID not set. Bot would respond to ALL users.")
        print("  Set TELEGRAM_USER_ID to your Telegram user ID (send /whoami to discover it).")
        print("  export TELEGRAM_USER_ID=$(powershell -Command \"[Environment]::GetEnvironmentVariable('TELEGRAM_USER_ID', 'User')\" | tr -d '\\r\\n')")
        sys.exit(1)

    if not os.environ.get("GROQ_API_KEY"):
        print("WARNING: GROQ_API_KEY not set. Voice input disabled.")

    app = ApplicationBuilder().token(token).build()
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("whoami", whoami_command))
    app.add_handler(CommandHandler("status", status_command))
    app.add_handler(CommandHandler("save", save_command))
    app.add_handler(CommandHandler("flag", flag_command))
    app.add_handler(CommandHandler("haiku", model_command))
    app.add_handler(CommandHandler("sonnet", model_command))
    app.add_handler(CommandHandler("opus", model_command))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("Dryft bot running. Send a message in Telegram.")
    proxy.print_status()
    app.run_polling()


if __name__ == "__main__":
    main()
